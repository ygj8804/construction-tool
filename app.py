import streamlit as st
import pandas as pd
import openpyxl
from io import BytesIO
from openpyxl.utils import column_index_from_string
from datetime import datetime, timedelta
from openpyxl.styles import Alignment

# [정제 함수] 눈에 보이지 않는 공백 및 특수문자 제거
def clean_text(text):
    if pd.isna(text): return ""
    return str(text).replace('\xa0', ' ').strip()

# 1. 페이지 설정
st.set_page_config(page_title="서원건설 단가 자동 입력기", layout="wide")

# 한국 시간(KST)으로 현재 시간을 가져오는 함수
def get_kst_now():
    return (datetime.utcnow() + timedelta(hours=9)).strftime('%Y-%m-%d %H:%M:%S')

# 세션 상태 초기화 (동기화 시간 저장용)
if 'sync_time' not in st.session_state:
    st.session_state.sync_time = get_kst_now()

# [데이터 연결 주소]
DATA_URL = "https://docs.google.com/spreadsheets/d/e/2PACX-1vT0RF-nXszGyvIIHGPfFJtgOvCnZrA_6A44Sq21te9CrOQuxYD_1Q5zO-9aZHLoHw/pub?gid=1069214405&single=true&output=csv"
EDIT_URL = "https://docs.google.com/spreadsheets/d/1XR0zYBVOL8PRJjuNvttpbo6WNH2fCRSt/edit?rtpof=true"

# 제목 및 상단 정보
st.title("🏗️ 서원건설 - 단가 자동 입력기")

# 3. 마스터 데이터 로드 (0값은 무시하고, 실제 값이 있는 경우만 가져오도록 수정)
@st.cache_data(ttl=60)
def load_master_data(duplicate_option):
    try:
        df = pd.read_csv(DATA_URL, header=None)
        temp_data = {} 
        
        for _, row in df.iterrows():
            # 키 생성
            key = f"{clean_text(row[0])}_{clean_text(row[1])}"
            
            # E(4), G(6), I(8) 행의 값을 문자열로 추출하여 정리
            e_str = str(row[4]).replace(',', '').strip() if pd.notna(row[4]) else ""
            g_str = str(row[6]).replace(',', '').strip() if pd.notna(row[6]) else ""
            i_str = str(row[8]).replace(',', '').strip() if pd.notna(row[8]) else ""
            
            # 숫자 변환
            try:
                e_val = float(e_str) if e_str != "" else 0
                g_val = float(g_str) if g_str != "" else 0
                i_val = float(i_str) if i_str != "" else 0
            except ValueError:
                continue 
            
            # E, G, I 모두 0이거나 비어있으면 이 행은 무시
            if e_val == 0 and g_val == 0 and i_val == 0:
                continue
                
            if key not in temp_data:
                temp_data[key] = {'E': [], 'G': [], 'I': []}
            
            # 0보다 큰 값만 리스트에 추가
            if e_val > 0: temp_data[key]['E'].append(e_val)
            if g_val > 0: temp_data[key]['G'].append(g_val)
            if i_val > 0: temp_data[key]['I'].append(i_val)
        
        master_data = {}
        for key, values in temp_data.items():
            # 리스트가 비어있으면 0 처리, 아니면 min/max 적용
            if duplicate_option == "최저가 적용":
                master_data[key] = {
                    'E': min(values['E']) if values['E'] else 0,
                    'G': min(values['G']) if values['G'] else 0,
                    'I': min(values['I']) if values['I'] else 0
                }
            else:
                master_data[key] = {
                    'E': max(values['E']) if values['E'] else 0,
                    'G': max(values['G']) if values['G'] else 0,
                    'I': max(values['I']) if values['I'] else 0
                }
        return master_data
    except Exception as e: 
        st.error(f"데이터 로드 실패: {e}")
        return None

# 상단에 데이터 동기화 시간 및 수정 버튼 배치
col_info, col_btn = st.columns([0.6, 0.4])

# 2. 정갈한 설정창
with st.expander("⚙️ [설정] 데이터 시작 행 및 각 항목별 위치 (열) 입력", expanded=True):
    st.subheader("1. 데이터 시작 행")
    start_row = st.number_input("데이터가 실제로 시작되는 행 번호", value=4, min_value=1)
    
    st.markdown("---")
    
    st.subheader("3. 중복 데이터 처리 방식")
    duplicate_option = st.selectbox("동일 품목/규격이 여러 개 있을 경우", ["최저가 적용", "최고가 적용"])

    st.markdown("---")
    
    st.subheader("2. 각 항목별 위치 (열) 입력 (알파벳 입력)")
    col1, col2 = st.columns(2)
    
    with col1:
        st.write("**기본 항목**")
        col_name_str = st.text_input("품명 (열)", value="A")
        col_spec_str = st.text_input("규격 (열)", value="B")
        col_unit_str = st.text_input("단위 (열)", value="C")
        
    with col2:
        st.write("**단가 데이터 입력 위치**")
        col_mat_str = st.text_input("재료비단가 (열)", value="E")
        col_lab_str = st.text_input("노무비단가 (열)", value="G")
        col_exp_str = st.text_input("경비단가 (열)", value="I")

# 데이터 로드 실행
master_data = load_master_data(duplicate_option)

with col_info:
    st.info(f"📋 마스터 데이터 기준 시간 (데이터동기화완료): {st.session_state.sync_time}")
    # 마스터 데이터 동기화 버튼 추가
    if st.button("🔄 마스터 데이터 강제 동기화"):
        st.cache_data.clear()
        st.session_state.sync_time = get_kst_now()
        st.rerun()
        
    if master_data:
        st.metric("총 마스터 데이터 항목", f"{len(master_data):,} 개")

with col_btn:
    st.link_button("단가 수정하기 ↗️", EDIT_URL)

st.markdown("---")

# 4. 파일 업로드 및 기능 실행
uploaded_file = st.file_uploader("작업할 견적서 엑셀 파일을 업로드하세요", type=['xlsx', 'xls', 'csv'])
file_ext = st.selectbox("저장할 파일 형식 선택", ["xlsx", "csv"])

ws = None
if uploaded_file and master_data:
    try:
        if uploaded_file.name.endswith('.csv'):
            st.warning("CSV 파일은 셀 서식이 유지되지 않을 수 있습니다.")
        else:
            try:
                wb = openpyxl.load_workbook(uploaded_file)
                selected_sheet = st.selectbox("작업할 시트를 선택하세요", wb.sheetnames)
                ws = wb[selected_sheet]
            except Exception as e:
                st.error("오류: 업로드한 파일이 유효한 .xlsx 형식이 아닙니다.")
                st.stop()

        if ws and st.button("단가 매칭 실행", type="primary"):
            c_name = column_index_from_string(col_name_str.strip().upper())
            c_spec = column_index_from_string(col_spec_str.strip().upper())
            c_mat = column_index_from_string(col_mat_str.strip().upper())
            c_lab = column_index_from_string(col_lab_str.strip().upper())
            c_exp = column_index_from_string(col_exp_str.strip().upper())

            count = 0
            cell_style = Alignment(horizontal='right', vertical='center') 
            number_fmt = '#,##0'

            for row in range(start_row, ws.max_row + 1):
                val_a = clean_text(ws.cell(row=row, column=c_name).value)
                val_b = clean_text(ws.cell(row=row, column=c_spec).value)
                key = f"{val_a}_{val_b}"
                
                if key in master_data:
                    # 0보다 큰 데이터가 있는 경우에만 입력 (0인 경우 입력 안 함)
                    if master_data[key]['E'] > 0:
                        ws.cell(row=row, column=c_mat).value = master_data[key]['E']
                        ws.cell(row=row, column=c_mat).number_format = number_fmt
                        ws.cell(row=row, column=c_mat).alignment = cell_style
                    
                    if master_data[key]['G'] > 0:
                        ws.cell(row=row, column=c_lab).value = master_data[key]['G']
                        ws.cell(row=row, column=c_lab).number_format = number_fmt
                        ws.cell(row=row, column=c_lab).alignment = cell_style
                    
                    if master_data[key]['I'] > 0:
                        ws.cell(row=row, column=c_exp).value = master_data[key]['I']
                        ws.cell(row=row, column=c_exp).number_format = number_fmt
                        ws.cell(row=row, column=c_exp).alignment = cell_style
                    
                    count += 1
            
            st.success(f"완료! 총 {count}개의 항목에 단가가 입력되었습니다.")
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            final_name = f"매칭완료_{uploaded_file.name.split('.')[0]}.{file_ext}"
            st.download_button(label=f"수정된 파일 다운로드 ({file_ext})", data=output, file_name=final_name)
            
    except Exception as e:
        st.error(f"오류: 설정한 열 위치가 올바른지 확인해주세요. 상세: {e}")

elif uploaded_file and not master_data:
    st.error("마스터 데이터를 불러오지 못했습니다.")
